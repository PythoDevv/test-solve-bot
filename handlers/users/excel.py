import os
import random

from aiogram import types
from aiogram.types import InputFile
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

from loader import bot, db, dp


async def write_data_chunk(wb, users_chunk, start_row):
    ws = wb.active
    for idx, user in enumerate(users_chunk, start=start_row):
        ws.append([idx, user[1], user[2], user[3]])


@dp.message_handler(text="Excel File")
async def marathon(message: types.Message):
    admins = await db.select_all_admins()
    admins_list = []
    for i in admins:
        admins_list.append(i[1])
    if message.from_user.id in admins_list:

        await message.answer(
            """"Ma'lumotlarni filega yozish boshlandi...\n\nBazada qancha ma'lumotlar ko'p bo'lsa shuncha ko'p vaqt ketadi.\n\nIltimos jarayon tugagunga qadar botdan foydalanman."""
        )

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "№"
        ws["B1"] = "To'liq ismi"
        ws["C1"] = "Telegram Username"
        ws["D1"] = "Telefon raqami"

        users = await db.select_all_users()
        chunk_size = 1000
        for i in range(0, len(users), chunk_size):
            chunk = users[i: i + chunk_size]
            await write_data_chunk(wb, chunk, start_row=i + 2)  # Start from row 2 (after headers)

        n = random.sample(range(1, 100), 1)
        m = random.sample(range(100, 1000), 1)

        file_name = f"Excel_{n}_{m}.xlsx"
        wb.save(file_name)

        file = InputFile(path_or_bytesio=file_name)
        await message.answer_document(document=file)

        os.remove(file_name)


async def excel():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "№"
    ws["B1"] = "To'liq ismi"
    ws["C1"] = "Telegram Username"
    ws["D1"] = "Telefon raqami"
    ws["E1"] = "Yoshi"
    ws["F1"] = "Viloyati"
    ws["G1"] = "Tumani"
    ws["H1"] = "Telegram Id"

    # Rows can also be appended
    userss = await db.select_top_users_list()
    counter = 0
    book_number = "-"
    is_none = ""
    for user in userss:
        counter += 1

        ws.append(
            [
                f"{counter}",
                f"{user[11]}",
                f"@{user[2]}",
                f"{user[3]}",
                f"{user[9]}",
                f"{user[8]}",
                f"{user[10]}",
                f"{user[6]}",
            ]
        )
    n = random.sample(range(1, 100), 1)
    m = random.sample(range(100, 1000), 1)

    wb.save(f"Excel_{n}_{m}.xlsx")
    file = InputFile(path_or_bytesio=f"Excel_{n}_{m}.xlsx")
    await bot.send_document(chat_id=935795577, document=file)
    os.remove(f"Excel_{n}_{m}.xlsx")


@dp.message_handler(text="read")
async def marathon(message: types.Message):
    book = load_workbook(filename="test_e.xlsx")
    ws = book.active

    for i in range(2, 15000):
        await db.add_usersss(
            name=ws[f"B{i}"].value,
            username=ws[f"C{i}"].value,
            phone=ws[f"D{i}"].value,
            year=ws[f"I{i}"].value,
            region=ws[f"F{i}"].value,
            district=ws[f"G{i}"].value,
            telegram_id=ws[f"H{i}"].value,
        )
